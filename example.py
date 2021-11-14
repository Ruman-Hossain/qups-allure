import openpyxl

from WhatsAppUIAutomation.automation import WhatsAppUi


def load_excel(r, input_message):
    workbook = openpyxl.load_workbook('./whatsapp_ui.xlsx')
    print(str(workbook) + " Workbook Opened...")
    sheet = workbook['whatsapp']
    print(str(sheet) + " Reading...")
    sl_no = sheet.cell(row=r, column=1).value
    number = sheet.cell(row=r, column=2).value

    whatsapp = WhatsAppUi()
    whatsapp.search_number(number)
    whatsapp.send_message(input_message)

    # OUTPUT WRITE TO EXCEL FILE
    sheet.cell(row=r, column=3).value = whatsapp.message
    sheet.cell(row=r, column=4).value = whatsapp.sent_status
    sheet.cell(row=r, column=5).value = whatsapp.read_status
    sheet.cell(row=r, column=6).value = whatsapp.login_status
    sheet.cell(row=r, column=7).value = whatsapp.logout_status
    workbook.save('./whatsapp_ui.xlsx')
    print("DATA INSERTED SUCCESSFULLY IN ROW " + str(r))
    # print(f'SL No : {sl_no} \n Number : {number} \n '
    #       f'Message : {message} \n Sent Status : {whatsapp.sent_status} \n '
    #       f'Checked : {whatsapp.read_status} \n '
    #       f'Login : {whatsapp.login_status} \n Logout : {whatsapp.logout_status}')


if __name__ == "__main__":

    while True:
        row = int(input('Enter Excel Row Number : '))
        message = input('Enter Message to Send : ')
        load_excel(row, message)

    # whatsapp = WhatsAppUi()
    # whatsapp.search_number("+8801402004389")
    # whatsapp.send_message("Message sent, Status Checked")
