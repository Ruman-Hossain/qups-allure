import openpyxl


class ExcelReadWrite(object):
    def __init__(self):
        self.workbook = openpyxl.load_workbook('../whatsapp_ui.xlsx')
        self.sheet = self.workbook['whatsapp']
        self.row_num = ''

    def read_number(self, row_num):
        return self.sheet.cell(row=row_num, column=1).value

    def write_output_in_excel(self, row_num, message, sent_status, read_status, login_status, logout_status):
        self.sheet.cell(row=row_num, column=3).value = message
        self.sheet.cell(row=row_num, column=4).value = sent_status
        self.sheet.cell(row=row_num, column=5).value = read_status
        self.sheet.cell(row=row_num, column=6).value = login_status
        self.sheet.cell(row=row_num, column=7).value = logout_status
        self.workbook.save('../whatsapp_ui.xlsx')
        print(f'Data Inserted Successfully in row {row_num}')
